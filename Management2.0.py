import os
import csv
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import design
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QTableWidget

class neededItems(object):
  
  def __init__(self,name=None,quantity=None,minimum=None):
        self.__quantity = int(quantity)
        self.name = name
        self.minimum = int(minimum)
  #Changing names and changing quantity of items
  def get_quantity(self):
    return str(self.__quantity)
  def get_minimum(self):
    return str(self.minimum)
  def changename(self,name):
    self.name = name
  def set_quantity(self, quantity):
        self.__quantity = int(quantity)
  def add_quantity(self, x):
        self.__quantity = self.__quantity + int(x)
  def minus_quantity(self, x):
        self.__quantity = self.__quantity - int(x)
  def set_minimum(self, x):
        self.minimum = int(x)

class inventoryItem(object):
  def __init__(self,assettag=None,serial=None):
    self.assettag = assettag
    self.serial= serial

WORKBOOK_NAME = "inventory.xlsx"
if (os.path.exists(WORKBOOK_NAME) == False):
  wb = openpyxl.Workbook()
  sheet = wb.active
  sheet.title= 'Sheet1'
  wb.save(WORKBOOK_NAME)
wb = openpyxl.load_workbook(WORKBOOK_NAME)
itemlist = []
currentindex = 0
def ifexists(name, itemlist):
  i = 0
  for neededitems in itemlist:
    if name in neededitems.name:
      return i
    i+=1
  return -1
                             
class ManagementApp(QtWidgets.QMainWindow, design.Ui_MainScreen):
    def __init__(self, parent=None):
        super(ManagementApp, self).__init__(parent)
        self.setupUi(self)
        self.setupapp()
    def setupapp(self):
      if (os.path.exists(WORKBOOK_NAME) == False):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title= 'Sheet1'
        wb.save(WORKBOOK_NAME)
      wb = openpyxl.load_workbook(WORKBOOK_NAME)
      sheet = wb['Sheet1']
      #Importing items from Spreadsheet if they exist
      if (sheet.max_row > 1):
        for i in range(2,sheet.max_row+1):
          item= neededItems(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=3).value)
          itemlist.append(item)
          self.comboBox.addItem(item.name)
      sheetlist = []
      sheetlist = wb.sheetnames
      itemnameslist= []
      lowitems = ""
      low = False
      for neededitems in itemlist:
          if int(neededitems.get_quantity()) < neededitems.minimum:
            lowitems = neededitems.name + ", "
            low = True
          itemnameslist.append(neededitems.name)
          if neededitems.name not in sheetlist:
            wb.create_sheet(neededitems.name)
            sheet= wb[neededitems.name]
            sheet['A1'] = 'Asset Tag'
            sheet['B1'] = 'Serial'
      for listname in sheetlist:
          if listname != 'Sheet1':
            if listname not in itemnameslist:
              wb.remove(wb[listname])
      lowitems= lowitems + "have low inventory"
      if (low== True):
        QMessageBox.question(self, "Low Inventory", lowitems, QMessageBox.Ok, QMessageBox.Ok)
      wb.save(WORKBOOK_NAME)
      neededitems = itemlist[currentindex]
      quantity = "Quantity: " + neededitems.get_quantity()
      minimum = "Minimum: " + neededitems.get_minimum()
      sheet2 = wb[neededitems.name]
      self.itemstable.setRowCount(sheet2.max_row)
      self.itemstable.setColumnCount(sheet2.max_column)
      for i in range(2,sheet2.max_row+1):
        for x in range(1,sheet2.max_column+1):
          value = sheet2.cell(row=i, column=x).value
          self.itemstable.setItem(i-2,x-1, QTableWidgetItem(value))
      self.label.setText(quantity)
      self.label_2.setText(minimum)
      self.currentitem.setText(neededitems.name)
      self.comboBox.activated[str].connect(self.changeCategory)
      self.ChangeQuantity.clicked.connect(self.changeQuantity)
      self.AddQuantity.clicked.connect(self.addQuantity)
      self.RemoveQuantity.clicked.connect(self.removeQuantity)
      self.Changeminimum.clicked.connect(self.changeMinimum)
      self.changename.clicked.connect(self.changeName)
      self.Addcategory.clicked.connect(self.addCategory)
      self.Removecategory.clicked.connect(self.removeCategory)
      self.save.clicked.connect(self.saveSheet)
      self.Additem.clicked.connect(self.addRow)
      self.Removeitem.clicked.connect(self.delRow)
      self.pushButton.clicked.connect(self.saveItem)
    def changeCategory(self, text):
      currentindex = ifexists(text,itemlist)
      neededitems = itemlist[currentindex]

      quantity = "Quantity: " + neededitems.get_quantity()
      
      minimum = "Minimum: " + neededitems.get_minimum()
      
      self.label.setText(quantity)
      self.label_2.setText(minimum)
      self.currentitem.setText(neededitems.name)
      sheet2 = wb[neededitems.name]
      self.itemstable.setRowCount(0)
      self.itemstable.setRowCount(sheet2.max_row)
      self.itemstable.setColumnCount(sheet2.max_column)
      for i in range(2,sheet2.max_row+1):
        for x in range(1,sheet2.max_column+1):
          
          value = sheet2.cell(row=i, column=x).value
          self.itemstable.setItem(i-2,x-1, QTableWidgetItem(value))
    def changeQuantity(self):
      textboxvalue = self.quantitinput.text()
      neededitems = itemlist[currentindex]
      neededitems.set_quantity(textboxvalue)
      quantity = "Quantity: " + textboxvalue
      self.label.setText(quantity)
    def addQuantity(self):
      textboxvalue = self.quantitinput.text()
      neededitems = itemlist[currentindex]
      neededitems.add_quantity(textboxvalue)
      quantity = "Quantity: " + neededitems.get_quantity()
      self.label.setText(quantity)
    def removeQuantity(self):
      textboxvalue = self.quantitinput.text()
      neededitems = itemlist[currentindex]
      neededitems.minus_quantity(textboxvalue)
      quantity = "Quantity: " + neededitems.get_quantity()
      self.label.setText(quantity)
    def changeMinimum(self):
      textboxvalue = self.quantitinput.text()
      neededitems = itemlist[currentindex]
      neededitems.set_minimum(textboxvalue)
      minimum = "Minimum: " + neededitems.get_minimum()
      self.label_2.setText(minimum)
    def changeName(self):
      textboxvalue = self.nameinput_2.text()
      neededitems = itemlist[currentindex]
      neededitems.name= textboxvalue
      self.comboBox.setItemText(currentindex,textboxvalue)
      self.currentitem.setText(textboxvalue)
    def addCategory(self):
      textboxvalue = self.nameinput.text()
      textboxvalue2 = self.quantityinput.text()
      newitem = neededItems(textboxvalue, textboxvalue2,5)
      itemlist.append(newitem)
      self.comboBox.addItem(textboxvalue)
      wb.create_sheet(title=textboxvalue)
      sheet= wb[textboxvalue]
      sheet['A1'] = 'Asset Tag'
      sheet['B1'] = 'Serial'
      wb.save(WORKBOOK_NAME)
    def removeCategory(self):
      neededitems = itemlist[currentindex]
      itemlist.remove(neededitems)
      self.comboBox.removeItem(currentindex)
      wb.remove(wb[neededitems.name])
      wb.save(WORKBOOK_NAME)
    def saveSheet(self):
      sheet = wb['Sheet1']
      sheet['A1'] = 'Name'
      sheet['B1'] = 'Quantity'
      sheet['C1'] = 'Minimum'
      i=2
      for neededitems in itemlist:
        x= neededitems.get_quantity()
        sheet['A' + str(i)]  = neededitems.name
        sheet['B'+ str(i)] = x
        sheet['C' + str(i)] = str(neededitems.minimum)
        i+=1
      wb.save(WORKBOOK_NAME)
      QMessageBox.question(self, "Saved Item", "Worksheet has been saved", QMessageBox.Ok, QMessageBox.Ok)
    def addRow(self):
      rowPosition = self.itemstable.rowCount()
      self.itemstable.insertRow(rowPosition)
    def delRow(self):
      index = self.itemstable.currentRow()
      self.itemstable.removeRow(index)
    def saveItem(self):
      neededitems = itemlist[currentindex]
      sheet2 = wb[neededitems.name]
      for i in range(0,self.itemstable.rowCount()):
        for x in range(0,self.itemstable.columnCount()):
          item = self.itemstable.item(i,x)
          if (item != None):
            value= item.text()
            sheet2.cell(row=i+2, column=x+1).value = value
            wb.save(WORKBOOK_NAME)
def main():
    app = QtWidgets.QApplication(sys.argv)
    form = ManagementApp()
    form.show()
    app.exec_()

if __name__ == '__main__':
    main()
